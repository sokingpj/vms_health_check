#!/usr/bin/python
from functools import wraps
import os
import re
from retrying import retry
import traceback
import openpyxl
from openpyxl.utils import get_column_letter

from .check_logger import Check_Logger
import argparse

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

class CollectResult():
    def __init__(self,resultFolder=None,resultXlsPath=None,logger=None) -> None:
        self.logger = logger
        if not self.logger:
            self.logger = Check_Logger(rootLogger=None,logger_name='')
        self.resultFolder = resultFolder
        if not os.path.exists(self.resultFolder):
            raise Exception('Please be sure the collect result folder is exist!!')
        self.indexs = ['Execute_Time','Check_Project','Correlation_Servers','HealthCheckRST','Check_Failed_Servers','Net_Area','CURR_PATH','Server_IP']
        '''01_ioc-01_ioc:
            10_250_143_11:
              Execute_Time: 20230217090547
              Check_Project: 01_ioc-01_ioc
              Correlation_Servers: 10.250.143.100,10.250.143.101
              HealthCheckRST: /home/healthCheck/health_checkRst_20230217090547.tgz
              Check_Failed_Servers: 10.250.143.100,10.250.143.101
              Health_Check_Item:
                app_docker_container_status: NotCheck
                os_os_cpu_usage: Success
            '''
        self.colRst = {}
        self.resultXlsPath = resultXlsPath

    def collect_RstFile(self,cmd: str=''):
        try:
            self.logger.info('process result folder: %s' % cmd)
            part = re.compile('[0-9]{1,3}_[0-9]{1,3}_[0-9]{1,3}_[0-9]{1,3}',re.M)
            self.fileList = {}
            for root,dir,files in os.walk(self.resultFolder):
                for file in files:
                    fileName = file.strip()
                    if fileName == 'check_items.conf':
                        ipDir = os.path.basename(root)
                        filePath = os.path.join(root,fileName)
                        comIPLst = part.findall(ipDir)
                        # self.logger.info('comIPLst: %s, ipDir: %s, type(ipDir): %s' % (comIPLst,ipDir, type(ipDir)))
                        if len(comIPLst) != 1:
                            continue
                        if ipDir not in self.fileList.keys():
                            self.fileList[ipDir] = filePath
            # self.logger.info(self.fileList)
        except Exception as e:
            raise Exception('Collect result files into a list has failed, exception %s, %s' % (e,traceback.format_exc()))
    
    def analysis_rstFile(self):
        try:
            if not self.fileList:
                raise Exception('self.fileList is None, skip to collect. self.fileList: %s' % (self.fileList))
            for ipDir,filePath in self.fileList.items():
                if not os.path.exists(filePath):
                    raise Exception('collect result file is not exist, please check!! filePath=%s' % filePath)
                
                proStr = os.path.basename(os.path.dirname(os.path.dirname(filePath)))
                if proStr not in self.colRst.keys():
                    self.colRst[proStr] = {}
                if ipDir not in self.colRst[proStr].keys():
                    self.colRst[proStr][ipDir] = {}
                
                if 'Health_Check_Item' not in self.colRst[proStr][ipDir].keys():
                    self.colRst[proStr][ipDir]['Health_Check_Item'] = {}
                frad = open(filePath, 'r', encoding='utf-8')
                lines = frad.readlines()
                for line in lines:
                    key = line.strip().split(' ')[0]
                    val = line.strip().split(' ')[-1]
                    if val == '|':
                        val = ''
                    if not line or line.startswith('----------------------------------------'):
                        continue
                    
                    if key in self.indexs:
                        if key not in self.colRst[proStr][ipDir].keys():
                            if key == 'Health_Check_Item':
                                self.colRst[proStr][ipDir][key] = {}
                            else:
                                self.colRst[proStr][ipDir][key] = val
                        else:
                            self.colRst[proStr][ipDir][key] = val
                    else:
                        if key == 'Health_Check_Item':
                            continue
                        if key not in self.colRst[proStr][ipDir]['Health_Check_Item'].keys():
                            self.colRst[proStr][ipDir]['Health_Check_Item'][key] = val
                frad.close()
            # self.logger.info(self.colRst)
        except Exception as e:
            raise Exception('analysis result files has failed, exception %s, %s' % (e,traceback.format_exc()))

    def save_to_excel(self):
        try:
            self.logger.info('start save_to_excel()...')
            if os.path.exists(self.resultXlsPath):
                os.remove(self.resultXlsPath)
            self.result_wb = None
            self.result_wb = openpyxl.Workbook()
            for proStr, vals in self.colRst.items():
                sheet_name = re.sub('[a-z_]','',proStr)
                for sName in ['Sheet', sheet_name]:
                    if sName in self.result_wb.sheetnames:
                        wSheet = self.result_wb[sName]
                        self.result_wb.remove(wSheet)
                rst_sheet = self.result_wb.create_sheet(title=sheet_name)
                
                indexs_len = len(self.indexs)
                ip0 = list(vals.keys())[0]
                r_index = 1
                for idx in self.indexs:
                    rst_sheet.cell(row=r_index, column=1).value = idx
                    rst_sheet.cell(row=r_index, column=2).value = vals[ip0][idx]
                    rst_sheet.column_dimensions[get_column_letter(1)].width = 21
                    r_index +=  1

                chk_itmes = list(vals[ip0]['Health_Check_Item'].keys())
                # self.logger.info('ip0: %s,chk_itmes: %s' % (ip0,chk_itmes))

                c_index = 2
                r_index = indexs_len + 2

                for item_k in chk_itmes:
                    item_col = c_index + chk_itmes.index(item_k)
                    rst_sheet.column_dimensions[get_column_letter(item_col)].width = len(item_k) + 2
                    rst_sheet.cell(row=r_index, column=item_col).value = item_k

                r_index = indexs_len + 3
                for kip,vals in vals.items():
                    rst_sheet.cell(row=r_index, column=1).value = kip.replace('_','.')
                    for item_k,item_v in vals['Health_Check_Item'].items():
                        item_col = c_index + chk_itmes.index(item_k)
                        rst_sheet.cell(row=r_index, column=item_col).value = item_v
                    r_index += 1

            self.logger.info('save health check result into: %s' % self.resultXlsPath)
            self.result_wb.save(self.resultXlsPath)   
        except Exception as e:
            raise Exception('Write cigarette data into %s has failed, error ino: %s.\n %s' %(self.resultXlsPath, e, traceback.format_exc()))

            
    def collectRst(self):
        try:
            self.collect_RstFile()
            self.analysis_rstFile()
            self.save_to_excel()
        except Exception as e: 
            raise Exception('collect check result has failed, error ino: %s.\n %s' %(e, traceback.format_exc()))

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='health check parameters.')
    parser.add_argument('--resultFolder', type=str, default='/tmp/health_check/result/20230217090928')
    parser.add_argument('--resultXlsPath', type=str, default='/tmp/health_check/result/healthCheck-20230217090928-03_urban_brain-03_industrial_economics-p2.xlsx')
    parser.add_argument('--logger', type=object, default=None, help='Should be logger object')
    args = parser.parse_args()

    print('collect health result from %s' % args.resultFolder)
    print('collect health result xlsx %s' % args.resultXlsPath)
    collectObj = CollectResult(resultFolder=args.resultFolder,resultXlsPath=args.resultXlsPath,logger=None)
    collectObj.collectRst()
