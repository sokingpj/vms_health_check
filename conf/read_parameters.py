#!/usr/bin/python
from functools import wraps
import os
import traceback
from yaml import loader
import yaml
import openpyxl
import base64
import hashlib
from .check_logger import Check_Logger


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
hashMd5 = hashlib.md5()

class Init_Parameters():
    def __init__(self,logger=None,lowSvrType=''):
        self.logger = logger
        if not self.logger:
            self.logger = Check_Logger(rootLogger=None,logger_name='')

        server_xlsx = "Servers-{lowSvrType}.xlsx".format(lowSvrType=lowSvrType)
        self.serversXlsPath = os.path.join(BASE_DIR, server_xlsx)

        self.config = os.path.join(BASE_DIR, 'config.yaml')
        self.init_config = os.path.join(BASE_DIR, 'init_config.yaml')

        self.logger.logger.info('%s: %s' % ('self.serversXlsPath', self.serversXlsPath))
        if not os.path.exists(self.serversXlsPath):
            self.logger.logger.error('The Servers xls file is not exist, please check!!')
            return

        self.servers_wb = None
        self.conf_E_Gov_yml = None
        self.svrsSheet = ['01_ioc', '02_digital_twins', '03_urban_brain', '04_shared_exchange']
        self.svr_pwd = {'exterior': {}, 'interior': {}}


    def function_logit(func):
        @wraps(func)
        def with_logging(*args, **kwargs):
            print('execute %s()...' % func.__name__)
            return func(*args, **kwargs)
        return with_logging


    @function_logit
    def read_config_yaml(self, yamlPath):
        try:
            read_yml_rst = None
            if not os.path.exists(yamlPath):
                raise Exception('The employee yaml file is not exist, please check!!')
            self.logger.logger.info('process yaml: %s' %yamlPath)
            fread = open(yamlPath, 'r', encoding='utf-8')
            read_yml_rst = yaml.load(fread.read(), Loader=loader.FullLoader)
            fread.close()
            return read_yml_rst
        except Exception as e:
            raise Exception('read_config_yaml has failed, %s, traceback.format_exc(): %s' % (e, traceback.format_exc()))


    @function_logit
    def __prepare_servers_wb(self):
        try:
            self.servers_wb = None
            if os.path.exists(self.serversXlsPath):
                self.servers_wb = openpyxl.load_workbook(self.serversXlsPath)
            if self.servers_wb == None:
                raise Exception('__prepare_servers_wb() read data has failed, self.servers_wb:%s, %s' %(self.servers_wb, traceback.format_exc()))
        except Exception as e:
            raise Exception('__prepare_servers_wb() read data from %s has failed, error ino: %s, %s' %(self.serversXlsPath, e, traceback.format_exc()))


    @function_logit
    def __read_svr_pwd(self,sheetName=''):
        '''
        read all servers's IP、user、password into the json data in the memory.
        data struct
        {
            projects:
            01_ioc:
                01_ioc:
                interior:
                    '10_250_143_100':
                    name: 城市运行综合管理中心软件_BPM应用
                    pwd: MVpoeGNxbCE=
                    user: root
        }
        '''
        try:
            if sheetName not in ['exterior','interior']:
                raise Exception('The server password sheet name should be "exterior" or "interior", please check.')
            sheets_list = self.servers_wb.get_sheet_names()
            if sheetName not in sheets_list:
                raise Exception('The sheet name(%s) was not in the xls file, please check. sheets in the xls file: %s' % (sheetName, sheets_list))
            sheetTable = self.servers_wb[sheetName]

            self.svr_pwd[sheetName] = {}
            for row in range(2, sheetTable.max_row, 1):
                svr_ip_tmp = sheetTable.cell(row, 2).value
                svr_usr = sheetTable.cell(row, 4).value
                svr_pwd = sheetTable.cell(row, 5).value
                if svr_ip_tmp is None or svr_usr is None or svr_pwd is None:
                    continue
                svr_ip = svr_ip_tmp.replace('.','_')
                # self.logger.logger.info('svr_ip: %s,svr_ip: %s,svr_ip: %s'%(svr_ip,svr_pwd,svr_usr))
                if svr_ip not in self.svr_pwd['exterior'].keys():
                    self.svr_pwd[sheetName][svr_ip] = {}
                self.svr_pwd[sheetName][svr_ip]['svr_usr'] = svr_usr
                self.svr_pwd[sheetName][svr_ip]['svr_pwd'] = str(base64.b64encode(svr_pwd.encode('utf-8')),encoding='utf-8')
                # self.svr_pwd[sheetName][svr_ip]['svr_pwd'] = svr_pwd

        except Exception as e:
            raise Exception('execute __read_child() has failed. {e} \n{traceback}'.format(e=e, traceback=traceback.format_exc()))


    @function_logit
    def __read_child(self, sheetName=''):
        try:
            if sheetName == '' or sheetName == None:
                raise Exception('The project name should not be None, please check.')
            sheets_list = self.servers_wb.get_sheet_names()
            if sheetName not in sheets_list:
                raise Exception('The sheet name(%s) was not in the xls file, please check. sheets in the xls file: %s' % (sheetName, sheets_list))
            sheetTable = self.servers_wb[sheetName]

            not_define = []
            for row in range(2, sheetTable.max_row, 1):
                child_tmp = sheetTable.cell(row, 1).value
                if child_tmp == '' or child_tmp == None:
                    continue
                
                child = child_tmp.lower()
                svr_type = sheetTable.cell(row, 2).value
                chinese_name = sheetTable.cell(row, 3).value
                svr_ip_tmp = sheetTable.cell(row, 4).value
                svr_ip = svr_ip_tmp.replace('.','_')
                # self.logger.logger.info('%10s, %10s, %10s, %10s, %10s' % (child, svr_type, chinese_name, svr_ip_tmp, sheetName))

                if child not in self.conf_E_Gov_yml['projects'][sheetName].keys():
                    self.conf_E_Gov_yml['projects'][sheetName][child] = {}
                    self.conf_E_Gov_yml['projects'][sheetName][child]['name'] = '未定义的子项'
                    self.conf_E_Gov_yml['projects'][sheetName][child]['interior'] = {}
                    self.conf_E_Gov_yml['projects'][sheetName][child]['exterior'] = {}

                svr_flag = 'interior'
                if svr_type.lower() == 'exterior':
                    svr_flag = 'exterior'
                if svr_ip not in self.conf_E_Gov_yml['projects'][sheetName][child][svr_flag].keys():
                    self.conf_E_Gov_yml['projects'][sheetName][child][svr_flag][svr_ip] = {}

                svr_usr = 'root'
                svr_pwd = ''
                if svr_ip in self.svr_pwd[svr_flag].keys():
                    svr_usr = self.svr_pwd[svr_flag][svr_ip]['svr_usr']
                    svr_pwd = self.svr_pwd[svr_flag][svr_ip]['svr_pwd']
                else:
                    not_define.append(svr_ip_tmp)

                self.conf_E_Gov_yml['projects'][sheetName][child][svr_flag][svr_ip]['name'] = chinese_name
                self.conf_E_Gov_yml['projects'][sheetName][child][svr_flag][svr_ip]['user'] = svr_usr
                self.conf_E_Gov_yml['projects'][sheetName][child][svr_flag][svr_ip]['pwd'] = svr_pwd
            self.logger.logger.info('未定义账户密码的虚机: %s, child: %s' % (not_define, sheetName))
            # self.logger.logger.info('%s' % self.conf_E_Gov_yml)
        except Exception as e:
            raise Exception('execute __read_child() has failed. {e} \n{traceback}'.format(e=e, traceback=traceback.format_exc()))

    @function_logit
    def write_into_yaml(self):
        try:
            self.logger.logger.info('\nServers info in: %s' % self.config)
            fwrite = open(self.config, 'w', encoding='utf-8')
            yaml.safe_dump(self.conf_E_Gov_yml, fwrite, sort_keys=True, encoding='utf-8', default_flow_style=False,
                        canonical=False, width=800, default_style=None, indent=2, allow_unicode=True)
            fwrite.close()

        except Exception as e:
            raise Exception('prepare Competence List has failed, error ino: %s, %s' %(e, traceback.format_exc()))

    @function_logit
    def get_conf_E_Gov_yml(self):
        self.conf_E_Gov_yml = self.read_config_yaml(self.init_config)
        self.__prepare_servers_wb()
        for item in ['exterior','interior']:
            self.__read_svr_pwd(item)
        for project in self.svrsSheet:
            self.__read_child(sheetName=project)
        self.write_into_yaml()
        return self.conf_E_Gov_yml

    @function_logit
    def book_main(self):
        try:
            self.conf_E_Gov_yml = self.read_config_yaml(self.init_config)
            self.__prepare_servers_wb()
            for item in ['exterior','interior']:
                self.__read_svr_pwd(item)
            for project in self.svrsSheet:
                self.__read_child(sheetName=project)
            self.write_into_yaml()
            return 0
        except Exception as e:
            self.logger.logger.error('init cigarette has failed, please check: {e} \n{traceback}'.format(e=e, traceback=traceback.format_exc()))
            return 1


if __name__ == "__main__":
    obj = Init_Parameters()
    ret = obj.book_main()
    exit(ret)
