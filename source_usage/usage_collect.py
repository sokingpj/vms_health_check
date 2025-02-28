#!/usr/bin/python
from functools import wraps
import os
import re
import tempfile
import tarfile
import traceback
import openpyxl
from openpyxl.utils import get_column_letter
import json
import calendar

from concurrent.futures import ThreadPoolExecutor,wait,ALL_COMPLETED
from check_logger import Check_Logger
import argparse
import shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

class CollectResult():
    def __init__(self,tarSAfilePath=None,resultXlsPath=None,logger=None) -> None:
        self.logger = logger
        if not self.logger:
            self.logger = Check_Logger(rootLogger=None,logger_name='')
        self.tarSAfilePath = tarSAfilePath
        self.rst_tmp_folder = tempfile.mkdtemp()
        self.year_month = ['2023','03']
        if not os.path.exists(self.tarSAfilePath):
            raise Exception('Please be sure the collect result file is exist!!')
        '''01_ioc-01_ioc:
            10_250_143_11:
              20230404:
                cpu: []
                mem: []
            '''
        self.colRst = {}
        self.resultXlsPath = os.path.join(BASE_DIR,os.path.basename(tarSAfilePath).split('.')[0] + '.xlsx')
        self.resultJsonPath = os.path.join(BASE_DIR,os.path.basename(tarSAfilePath).split('.')[0] + '.json')

    def prepare_day_list(self):
        self.month_day_list = []
        monthRange = calendar.monthrange(int(self.year_month[0]),int(self.year_month[1]))
        strTime = ''.join(self.year_month)
        for i in range(1,monthRange[-1]+1,1):
            self.month_day_list.append(strTime + (2-len(str(i)))*'0'+str(i))

    def decompress_tgzfiles(self,tmp_folder='',tarSAfilePath=''):
        try:
            if not tmp_folder or not tarSAfilePath:
                raise Exception('tmp_folder or tarSAfilePath should not be empty, tmp_folder=%s,tmp_folder=%s' % (tmp_folder,tarSAfilePath))
            
            if not os.path.exists(tmp_folder):
                os.makedirs(tmp_folder)

            self.logger.info('def decompress_tgzfiles({tar_file_path},{tmp_folder}'.format(tar_file_path=tarSAfilePath,tmp_folder=tmp_folder))
            if not os.path.exists(tarSAfilePath):
                raise Exception('tar file is not exist, please check.')
            
            tarf = tarfile.open(tarSAfilePath,'r:gz')            
            file_names = tarf.getnames()
            for file_name in file_names:
                tarf.extract(file_name, tmp_folder)
            tarf.close()
            self.logger.info('tar xf %s into %s' % (tarSAfilePath,tmp_folder))
        except Exception as e:
            self.logger.error('tar file has failed, exception %s, trace: %s'%(e, traceback.format_exc()))
        
    def collect_RstFile(self):
        try:
            self.logger.info('collect_RstFile()...')
            part = re.compile('[0-9]{1,3}_[0-9]{1,3}_[0-9]{1,3}_[0-9]{1,3}',re.M)
            self.fileList = {}
            resultFolder = os.path.join(self.rst_tmp_folder,'result')
            for root,dir,files in os.walk(resultFolder):
                for file in files:
                    fileName = file.strip()
                    if fileName.startswith('health_checkRst_'):
                        ipDir = os.path.basename(root)
                        filePath = os.path.join(root,fileName)
                        comIPLst = part.findall(ipDir)
                        # self.logger.info('comIPLst: %s, ipDir: %s, type(ipDir): %s' % (comIPLst,ipDir, type(ipDir)))
                        if len(comIPLst) != 1:
                            continue
                        if ipDir not in self.fileList.keys():
                            self.fileList[ipDir] = filePath
            # self.logger.info(self.fileList)
        except Exception as e:
            raise Exception('Collect result files into a list has failed, exception %s, %s' % (e,traceback.format_exc()))
    
    def do_analysis_saFile(self,proStr='',ipDir='',sa_file=''):
        try:
            # sa_file = r'D:\soche\Downloads\tmpsy0rcs84\sar03'
            if not os.path.exists(sa_file):
                raise Exception('self.sa_folder is None, skip to collect. sa_file: %s' % (sa_file))

            self.logger.info('process sar: %s' %(sa_file))
            cpu1_compile = re.compile('[ |\t]+.*CPU[ |\t]+.*idle',re.M)
            cpu2_compile = re.compile('[0-9]{2}.*[ |\t]+.*CPU[ |\t]+.*%idle',re.M)
            mem_compile = re.compile('kbmemfree[ |\t]+.*kbmemused[ |\t]+.*kbbuffers',re.M)
            lix_compile = re.compile('Linux.*CPU',re.M)
            # cpu: Average:        CPU      %usr     %nice      %sys   %iowait    %steal      %irq     %soft    %guest    %gnice     %idle
            # mem: 00:00:01    kbmemfree   kbavail kbmemused  %memused kbbuffers  kbcached  kbcommit   %commit  kbactive   kbinact   kbdirty  kbanonpg    kbslab  kbkstack   kbpgtbl  kbvmused
            # 取此段的最末尾行            
            tmp_dict = {"cpu":0,"mem":0}
            cpu_flag = False
            mem_flag = False
            cpu_idle_list = []
            memused_idx = 3
            
            day_time = None
            file = open(sa_file,'r',encoding='utf-8')            # open file
            file_data = file.readlines()        # read all file lines
            for rowTmp in file_data:
                row = rowTmp.strip()
                if not day_time:
                    lix_result = lix_compile.findall(row)
                    if lix_result:
                        day_time = row.split()[3].replace('-','')
                        # self.logger.info('row: %s, day_time: %s' %(row,day_time))

                if not cpu_flag:
                    cpu1_result = cpu1_compile.findall(row)
                    cpu2_result = cpu2_compile.findall(row)
                    if cpu1_result and not cpu2_result:
                        cpu_flag = True
                        continue
                
                # self.logger.info('row:%s, cpu1_result:%s,cpu2_result:%s' %(row,cpu1_result,cpu2_result))
                if cpu_flag and not row:
                    cpu_flag = False
                if cpu_flag:
                    cpu_idle_list.append(float(row.split()[-1]))
                
                if not mem_flag:
                    mem_result = mem_compile.findall(row)
                    # self.logger.info('mem_result: %s,row: %s' %(mem_result,row))
                    if mem_result:
                        mem_flag = True
                        memused_split = row.split()
                        memused_idx = memused_split.index('%memused')
                        continue

                # self.logger.info('memused_idx: %s' % memused_idx)
                if mem_flag and (row.startswith('平均时间:') or row.startswith('Average:')):
                    row_split = row.split()
                    tmp_dict['mem'] = '{}%'.format(row_split[memused_idx])
                    mem_flag = False
            
            cpu_tmp = '{:.2f}%'.format(100 - (sum(cpu_idle_list)/len(cpu_idle_list)))
            tmp_dict['cpu'] = cpu_tmp
            if day_time not in self.colRst[proStr][ipDir].keys():
                self.colRst[proStr][ipDir][day_time] = tmp_dict
            # self.logger.info('tmp_dict: %s' % (tmp_dict))
            
        except Exception as e:
            raise Exception('do analysis sa result files has failed, exception %s, %s' % (e,traceback.format_exc()))

    def do_analysis_sarTgzFile(self,proStr='',ipDir='',filePath=''):
        try:
            sa_folder = tempfile.mkdtemp(prefix='do_analysis_sarTgzFile')
            self.decompress_tgzfiles(sa_folder,filePath)
            if not os.path.exists(sa_folder):
                raise Exception('self.sa_folder is None, skip to collect. sa_folder: %s' % (sa_folder))
            # sa_folder = r'C:\Users\soche\AppData\Local\Temp\tmpsy0rcs84'

            file_list = []
            for root,dir,files in os.walk(sa_folder):
                for file in files:
                    fileName = file.strip()
                    if fileName.startswith('sar'):
                        file_list.append(os.path.join(root,file))
                        
            exec = ThreadPoolExecutor(max_workers=50)
            all_tasks = []
            for sarTgz in file_list:
                all_tasks.append(exec.submit(self.do_analysis_saFile, proStr=proStr,ipDir=ipDir,sa_file=sarTgz))
            self.logger.info(all_tasks)
            wait(all_tasks,return_when=ALL_COMPLETED)

            if os.path.exists(sa_folder):
                shutil.rmtree(sa_folder)
            self.logger.info('collect all sar file task has completed.')
        except Exception as e:
            raise Exception('do analysis sa result files has failed, exception %s, %s' % (e,traceback.format_exc()))

    # prepare the sarXX tar.gz file to analysis
    def analysis_sarFile(self):
        try:
            if not self.fileList:
                raise Exception('self.fileList is None, skip to collect. self.fileList: %s' % (self.fileList))
            all_tasks = []
            exec = ThreadPoolExecutor(max_workers=50)
            for ipDir,filePath in self.fileList.items():
                if not os.path.exists(filePath):
                    raise Exception('collect result file is not exist, please check!! filePath=%s' % filePath)
                
                # proStr = 04_shared_exchange-01_shared_exchange
                proStr = os.path.basename(os.path.dirname(os.path.dirname(filePath)))
                if proStr not in self.colRst.keys():
                    self.colRst[proStr] = {}
                # ipDir = 10_250_10_75
                if ipDir not in self.colRst[proStr].keys():
                    self.colRst[proStr][ipDir] = {}

                # self.logger.info('proStr=%s,ipDir=%s' %(proStr,ipDir))
                all_tasks.append(exec.submit(self.do_analysis_sarTgzFile, proStr=proStr,ipDir=ipDir,filePath=filePath))

            self.logger.info(all_tasks)
            wait(all_tasks,return_when=ALL_COMPLETED)
                
        except Exception as e:
            raise Exception('analysis result files has failed, exception %s, %s' % (e,traceback.format_exc()))
            
        
    def save_to_excel(self):
        try:
            self.logger.info('start save_to_excel()...')
            if os.path.exists(self.resultXlsPath):
                os.remove(self.resultXlsPath)
            self.result_wb = None
            self.result_wb = openpyxl.Workbook()
            for proStr, vals in self.colRst.items():
                sheet_name = re.sub('[a-z_]','',proStr)
                for sName in ['Sheet', sheet_name]:
                    if sName in self.result_wb.sheetnames:
                        wSheet = self.result_wb[sName]
                        self.result_wb.remove(wSheet)
                rst_sheet = self.result_wb.create_sheet(title=sheet_name)
                
                rst_sheet.cell(row=1, column=1).value = proStr
                chk_iplist = sorted(list(vals.keys()))
                col_idx = 2
                for ipStr in chk_iplist:
                    rst_sheet.cell(row=1, column=col_idx).value = ipStr
                    rst_sheet.cell(row=2, column=col_idx).value = 'CPU'
                    rst_sheet.cell(row=2, column=col_idx+1).value = 'Memory'
                    col_idx += 2

                rst_sheet.column_dimensions[get_column_letter(1)].width = 10
                r_idx = 4
                for idx in range(0,len(self.month_day_list),1):
                    row = r_idx + idx
                    day = self.month_day_list[idx]
                    rst_sheet.cell(row=row, column=1).value = day
                    
                col_idx = 2
                row_idx = 0
                for ipStr in chk_iplist:
                    chk_daylist = vals[ipStr].keys()
                    for day in self.month_day_list:
                        if day in chk_daylist:
                            day_idx = self.month_day_list.index(day)
                            row_idx = day_idx + r_idx
                            rst_sheet.cell(row=row_idx, column=col_idx).value = vals[ipStr][day]['cpu']
                            rst_sheet.cell(row=row_idx, column=col_idx+1).value = vals[ipStr][day]['mem']
                    col_idx += 2
                
                max_row = row_idx + 1
                max_column = len(chk_iplist) + 1
                rst_sheet.cell(row=max_row, column=1).value = 'AVERAGE'
                for col_tmp in range(1,max_column,1):
                    col = col_tmp * 2
                    col_letter1 = get_column_letter(col)
                    col_letter2 = get_column_letter(col+1)
                    rst_sheet.cell(row=max_row, column=col).value = '=AVERAGE({col_letter}4:{col_letter}{max_col})'.format(col_letter=col_letter1,max_col=row_idx)
                    rst_sheet.cell(row=max_row, column=col+1).value = '=AVERAGE({col_letter}4:{col_letter}{max_col})'.format(col_letter=col_letter2,max_col=row_idx)

            self.logger.info('save CPU amd Memory usage result into: %s' % self.resultXlsPath)
            self.result_wb.save(self.resultXlsPath)   
        except Exception as e:
            raise Exception('Write cigarette data into %s has failed, error ino: %s.\n %s' %(self.resultXlsPath, e, traceback.format_exc()))
        
    def save_json_result(self):
        try:
            self.logger.info('save json data into %s' % (self.resultJsonPath))
            if os.path.exists(self.resultJsonPath):
                os.remove(self.resultJsonPath)
            with open(self.resultJsonPath,'w',encoding='utf-8') as file:
                file.write(json.dumps(self.colRst, indent=2, ensure_ascii=False))
        except Exception as e:
            raise Exception('save collect result into %s has failed, error ino: %s.\n %s' %(e, traceback.format_exc()))
            
    def collectRst(self):
        try:
            self.prepare_day_list()
            self.decompress_tgzfiles(self.rst_tmp_folder,self.tarSAfilePath)
            self.collect_RstFile()
            self.analysis_sarFile()
            if os.path.exists(self.rst_tmp_folder):
                shutil.rmtree(self.rst_tmp_folder)
            
            self.save_json_result()            
            # self.logger.info('--------------------------')        
            # self.logger.info(self.colRst)
            self.save_to_excel()
        except Exception as e: 
            raise Exception('collect check result has failed, error ino: %s.\n %s' %(e, traceback.format_exc()))

# tarSAfilePath=None,resultXlsPath
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='health check parameters.')
    parser.add_argument('--tarSAfilePath', type=str, default='D:/01.CEC/02.Coding/02.usage/data/arm-gov-20230404091341.tgz')
    args = parser.parse_args()

    collectObj = CollectResult(tarSAfilePath=args.tarSAfilePath)
    collectObj.collectRst()
